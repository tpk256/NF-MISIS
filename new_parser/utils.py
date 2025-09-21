import hashlib
import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def file_hash(path: str, chunk_size: int = 8192) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            h.update(chunk)
    return h.hexdigest()


def get_current_parity() -> int:
    current_week = datetime.date.today().isocalendar()[1]  # Получаем текущую неделю
    return current_week % 2


def get_parity_from_excel(file_path: str) -> int:
    wb = load_workbook(file_path)

    flag_parity_odd = False
    flag_parity_even = False

    for work_sheet in wb.worksheets:
        if flag_parity_odd or flag_parity_even:
            break

        if find_cell_corners(work_sheet, "Нечетная неделя")[0]:
            flag_parity_odd = True

        if find_cell_corners(work_sheet, "Четная неделя")[0]:
            flag_parity_even = True

    if flag_parity_odd:
        return 1

    if flag_parity_even:
        return 0

    raise ValueError("Не смог определить четность")




def find_cell_corners(ws, target_value, flag_in=False):
    """
        Ищет в ws (Worksheet) ячейку со значением target_value.
        Если она в составе merged cell, возвращает углы этого диапазона,
        иначе – просто её координату дважды.
        Возвращает кортеж (top_left, bottom_right) как строки вида 'B3'.
    """
    # Сначала находим саму ячейку
    for row in ws.iter_rows():
        for cell in row:
            if target_value in str(cell.value):

                if not flag_in:
                    if target_value != cell.value:
                        continue
                coord = cell.coordinate

                # Проверяем, в какой merged-диапазон она входит
                for merge_range in ws.merged_cells.ranges:

                    if coord in merge_range:  # например 'B3' in 'B3:D5'
                        # bounds = (min_col, min_row, max_col, max_row)
                        min_col, min_row, max_col, max_row = merge_range.bounds
                        top_left = f"{get_column_letter(min_col)}{min_row}"
                        bottom_right = f"{get_column_letter(max_col)}{max_row}"
                        return top_left, bottom_right

                # Если не в merged, то углы – она сама
                return coord, coord

    # Если ничего не найдено
    return None, None
